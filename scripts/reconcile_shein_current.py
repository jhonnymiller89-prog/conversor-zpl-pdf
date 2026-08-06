from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pdfplumber


ROOT = Path("/Users/jhonnymiller/Documents/programa decodificador")
SHEIN_PATH = Path("/Users/jhonnymiller/Downloads/2438356_20260729002019.xlsx")
BANK_PDF_PATH = Path("/Users/jhonnymiller/Downloads/Bradesco_28072026_131839.PDF")
OUT_DIR = ROOT / "outputs/conciliacao_shein_20260728"
OUT_JSON = OUT_DIR / "conciliacao_shein_20260728.json"

DATE_RE = re.compile(r"^(\d{2}/\d{2}/\d{4})\b")
MONEY_RE = re.compile(r"(-?[\d.]+,\d{2})\s+(-?[\d.]+,\d{2})$")


def parse_datetime(value: Any) -> datetime | None:
    if value in (None, "-"):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                pass
    return None


def br_to_float(value: str | int | float | None) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    return round(float(value.replace(".", "").replace(",", ".")), 2)


def money(value: Any) -> float:
    return round(float(value or 0), 2)


def is_tx_line(line: str) -> bool:
    return bool(MONEY_RE.search(line))


def load_shein() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(SHEIN_PATH, data_only=True)
    ws = wb["sheet1"]
    headers = [cell.value for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    current_group = None
    current_total = None
    for raw in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, raw))
        if record.get("número da transação"):
            current_group = record["número da transação"]
            current_total = record["Valor total"]
        record["Grupo retirada"] = current_group
        record["Valor total do grupo"] = money(current_total)
        record["Valor recebido"] = money(record.get("Valor recebido"))
        record["valor líquido"] = money(record.get("valor líquido"))
        record["taxa de manuseio"] = money(record.get("taxa de manuseio"))
        record["Tempo de retirada_dt"] = parse_datetime(record.get("Tempo de retirada"))
        record["Tempo de sucesso_dt"] = parse_datetime(record.get("Tempo de sucesso de retirada"))
        rows.append(record)
    return rows


def load_bank_credits() -> list[dict[str, Any]]:
    with pdfplumber.open(str(BANK_PDF_PATH)) as pdf:
        lines: list[tuple[int, str]] = []
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            for line in text.splitlines():
                line = line.strip()
                if line:
                    lines.append((page_num, line))

    credits: list[dict[str, Any]] = []
    current_date = ""
    for idx, (page_num, line) in enumerate(lines):
        m_date = DATE_RE.match(line)
        if m_date:
            current_date = m_date.group(1)
        m_money = MONEY_RE.search(line)
        if not m_money:
            continue
        value_txt, balance_txt = m_money.groups()
        value = br_to_float(value_txt)
        if value <= 0:
            continue
        prefix = line[: m_money.start()].strip()
        tx_date = current_date
        m_inline_date = DATE_RE.match(prefix)
        if m_inline_date:
            tx_date = m_inline_date.group(1)
            prefix = prefix[m_inline_date.end() :].strip()

        parts = prefix.split()
        doc = ""
        doc_pos = None
        for pos in range(len(parts) - 1, -1, -1):
            if parts[pos].isdigit():
                doc = parts[pos]
                doc_pos = pos
                break
        if doc_pos is None:
            continue
        launch = " ".join(parts[:doc_pos]).strip()
        if not launch and idx > 0:
            previous = lines[idx - 1][1]
            if not is_tx_line(previous) and not previous.startswith(("REM:", "REMET.", "DES:", "Data ")):
                launch = previous
        detail = ""
        if idx + 1 < len(lines):
            next_line = lines[idx + 1][1]
            if not is_tx_line(next_line) and not next_line.startswith(("Data ", "Total ")):
                detail = next_line
        credits.append(
            {
                "Data": tx_date,
                "Data_dt": parse_datetime(tx_date),
                "Lancamento": launch,
                "Documento": doc,
                "Credito": value,
                "Saldo apos": br_to_float(balance_txt),
                "Detalhe": detail,
                "Linha original": line,
                "Pagina": page_num,
            }
        )
    return credits


def reconcile(shein_rows: list[dict[str, Any]], bank_credits: list[dict[str, Any]]) -> dict[str, Any]:
    credits_by_value: dict[float, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for idx, credit in enumerate(bank_credits):
        credits_by_value[credit["Credito"]].append((idx, credit))

    matched_bank: set[int] = set()
    reconciled: list[dict[str, Any]] = []
    pending: list[dict[str, Any]] = []
    shein_success = [
        row for row in shein_rows if row.get("Status de retirada") == "Retirada bem -sucedida" and row["Valor recebido"] > 0
    ]
    shein_success.sort(key=lambda r: (r["Tempo de sucesso_dt"] or datetime.min, r["Valor recebido"]))

    for row in shein_success:
        value = row["Valor recebido"]
        candidates = []
        for idx, credit in credits_by_value.get(value, []):
            if idx in matched_bank:
                continue
            shein_dt = row["Tempo de sucesso_dt"] or row["Tempo de retirada_dt"]
            bank_dt = credit["Data_dt"]
            days = abs((bank_dt.date() - shein_dt.date()).days) if bank_dt and shein_dt else 9999
            candidates.append((days, idx, credit))
        if candidates:
            days, idx, credit = sorted(candidates, key=lambda item: (item[0], item[1]))[0]
            matched_bank.add(idx)
            reconciled.append(
                {
                    "Grupo retirada": row["Grupo retirada"],
                    "Pedido retirada": row["Número do pedido de retirada"],
                    "Tempo retirada": row["Tempo de retirada_dt"],
                    "Tempo sucesso": row["Tempo de sucesso_dt"],
                    "Valor Shein": value,
                    "Status Shein": row["Status de retirada"],
                    "Data banco": credit["Data_dt"],
                    "Lancamento banco": credit["Lancamento"],
                    "Documento banco": credit["Documento"],
                    "Detalhe banco": credit["Detalhe"],
                    "Credito banco": credit["Credito"],
                    "Saldo banco": credit["Saldo apos"],
                    "Diferenca": round(credit["Credito"] - value, 2),
                    "Dias entre datas": days,
                    "Situacao": "Conciliado por valor exato",
                }
            )
        else:
            pending.append(
                {
                    "Grupo retirada": row["Grupo retirada"],
                    "Pedido retirada": row["Número do pedido de retirada"],
                    "Tempo retirada": row["Tempo de retirada_dt"],
                    "Tempo sucesso": row["Tempo de sucesso_dt"],
                    "Valor Shein": value,
                    "Status Shein": row["Status de retirada"],
                    "Situacao": "A receber - nao localizado no extrato",
                }
            )

    bank_unmatched = []
    for idx, credit in enumerate(bank_credits):
        if idx not in matched_bank:
            bank_unmatched.append({**credit, "Situacao": "Credito bancario sem item Shein correspondente"})

    return {
        "shein_rows": shein_rows,
        "bank_credits": bank_credits,
        "reconciled": reconciled,
        "pending": pending,
        "bank_unmatched": bank_unmatched,
        "summary": {
            "shein_count": len(shein_success),
            "shein_total": round(sum(row["Valor recebido"] for row in shein_success), 2),
            "reconciled_count": len(reconciled),
            "reconciled_total": round(sum(row["Valor Shein"] for row in reconciled), 2),
            "pending_count": len(pending),
            "pending_total": round(sum(row["Valor Shein"] for row in pending), 2),
            "bank_credit_count": len(bank_credits),
            "bank_credit_total": round(sum(row["Credito"] for row in bank_credits), 2),
            "bank_unmatched_count": len(bank_unmatched),
            "bank_unmatched_total": round(sum(row["Credito"] for row in bank_unmatched), 2),
        },
    }


def json_default(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


if __name__ == "__main__":
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    shein = load_shein()
    bank = load_bank_credits()
    result = reconcile(shein, bank)
    OUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    print(json.dumps(result["summary"], ensure_ascii=False, indent=2))
    for row in result["reconciled"]:
        print("OK", row["Pedido retirada"], row["Valor Shein"], row["Data banco"].date(), row["Documento banco"], row["Detalhe banco"])
    for row in result["pending"]:
        print("PEND", row["Pedido retirada"], row["Valor Shein"], row["Tempo sucesso"].date())
