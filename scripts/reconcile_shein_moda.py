from __future__ import annotations

from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path("/Users/jhonnymiller/Documents/programa decodificador")
SHEIN_PATH = Path("/Users/jhonnymiller/Downloads/financeiro shein mes 5 a 7 .xlsx")
BANK_PATH = ROOT / "output/transacoes_moda_mundial_brasil.xlsx"
OUT_DIR = ROOT / "outputs/shein_moda_conciliacao"
ANALYSIS_JSON = OUT_DIR / "reconciliation_data.json"


def parse_datetime(value: Any) -> datetime | None:
    if value in (None, "-"):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                pass
    return None


def money(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return round(float(value), 2)


def load_shein_rows() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(SHEIN_PATH, data_only=True)
    ws = wb["sheet1"]
    headers = [cell.value for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    current_transaction = None
    current_total = None
    for raw in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, raw))
        if record["número da transação"]:
            current_transaction = record["número da transação"]
            current_total = record["Valor total"]
        record["Grupo retirada"] = current_transaction
        record["Valor total do grupo"] = money(current_total)
        record["Valor recebido"] = money(record["Valor recebido"])
        record["valor líquido"] = money(record["valor líquido"])
        record["Tempo de retirada_dt"] = parse_datetime(record["Tempo de retirada"])
        record["Tempo de sucesso_dt"] = parse_datetime(record["Tempo de sucesso de retirada"])
        rows.append(record)
    return rows


def load_bank_rows() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(BANK_PATH, data_only=True)
    ws = wb["Transacoes"]
    headers = [cell.value for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, raw))
        record["Data_dt"] = parse_datetime(record["Data"])
        record["Credito"] = money(record["Credito"])
        record["Saldo apos lancamento"] = money(record["Saldo apos lancamento"])
        rows.append(record)
    return rows


def summarize_shein(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    details_by_group: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        group = row["Grupo retirada"] or "(sem grupo)"
        details_by_group[group].append(row)
        if group not in groups:
            groups[group] = {
                "Grupo retirada": group,
                "Valor total": row["Valor total do grupo"],
                "Tempo de retirada": row["Tempo de retirada_dt"],
                "Itens": 0,
                "Itens bem-sucedidos": 0,
                "Itens pendentes": 0,
                "Valor bem-sucedido": 0.0,
                "Valor pendente": 0.0,
                "Primeiro sucesso": None,
                "Ultimo sucesso": None,
                "Status": set(),
                "Pedidos": [],
            }
        g = groups[group]
        g["Itens"] += 1
        g["Status"].add(row["Status de retirada"])
        g["Pedidos"].append(row["Número do pedido de retirada"])
        if row["Status de retirada"] == "Retirada bem -sucedida":
            g["Itens bem-sucedidos"] += 1
            g["Valor bem-sucedido"] = round(g["Valor bem-sucedido"] + row["Valor recebido"], 2)
            success_dt = row["Tempo de sucesso_dt"]
            if success_dt and (g["Primeiro sucesso"] is None or success_dt < g["Primeiro sucesso"]):
                g["Primeiro sucesso"] = success_dt
            if success_dt and (g["Ultimo sucesso"] is None or success_dt > g["Ultimo sucesso"]):
                g["Ultimo sucesso"] = success_dt
        else:
            g["Itens pendentes"] += 1
            g["Valor pendente"] = round(g["Valor pendente"] + row["Valor recebido"], 2)
    output = []
    for group, g in groups.items():
        g["Status"] = ", ".join(sorted(s for s in g["Status"] if s))
        g["Pedidos"] = ", ".join(str(p) for p in g["Pedidos"] if p)
        output.append(g)
    output.sort(key=lambda r: r["Tempo de retirada"] or datetime.min)
    return output


def reconcile(
    shein_items: list[dict[str, Any]],
    bank_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    matched_bank_indexes: set[int] = set()
    reconciled: list[dict[str, Any]] = []
    shein_pending: list[dict[str, Any]] = []

    candidates_shein = [
        row for row in shein_items if row["Status de retirada"] == "Retirada bem -sucedida" and row["Valor recebido"] > 0
    ]
    candidates_shein.sort(
        key=lambda r: (
            r["Tempo de sucesso_dt"] or r["Tempo de retirada_dt"] or datetime.min,
            r["Valor recebido"],
            str(r["Número do pedido de retirada"]),
        )
    )

    for row in candidates_shein:
        value = row["Valor recebido"]
        candidates = []
        for idx, bank in enumerate(bank_rows):
            if idx in matched_bank_indexes:
                continue
            diff = round(bank["Credito"] - value, 2)
            if abs(diff) <= 0.01:
                bank_date = bank["Data_dt"]
                ref_date = row["Tempo de sucesso_dt"] or row["Tempo de retirada_dt"]
                days = abs((bank_date.date() - ref_date.date()).days) if bank_date and ref_date else 9999
                candidates.append((days, idx, bank, diff))
        if candidates and value > 0:
            days, idx, bank, diff = sorted(candidates, key=lambda x: (x[0], x[1]))[0]
            matched_bank_indexes.add(idx)
            reconciled.append(
                {
                    "Grupo retirada": row["Grupo retirada"],
                    "Número do pedido de retirada": row["Número do pedido de retirada"],
                    "Tempo de retirada": row["Tempo de retirada_dt"],
                    "Tempo de sucesso": row["Tempo de sucesso_dt"],
                    "Status Shein": row["Status de retirada"],
                    "Valor Shein": row["Valor recebido"],
                    "Valor total do grupo": row["Valor total do grupo"],
                    "Data banco": bank["Data_dt"],
                    "Documento banco": bank["Documento"],
                    "Credito banco": bank["Credito"],
                    "Saldo banco": bank["Saldo apos lancamento"],
                    "Diferenca": diff,
                    "Dias entre retirada e banco": days,
                    "Situacao conciliacao": "Conciliado",
                }
            )
        else:
            shein_pending.append(
                {
                    "Grupo retirada": row["Grupo retirada"],
                    "Número do pedido de retirada": row["Número do pedido de retirada"],
                    "Tempo de retirada": row["Tempo de retirada_dt"],
                    "Tempo de sucesso": row["Tempo de sucesso_dt"],
                    "Status Shein": row["Status de retirada"],
                    "Valor Shein": row["Valor recebido"],
                    "Valor total do grupo": row["Valor total do grupo"],
                    "Situacao conciliacao": "Nao localizado no banco",
                }
            )

    for row in shein_items:
        if row["Status de retirada"] != "Retirada bem -sucedida":
            shein_pending.append(
                {
                    "Grupo retirada": row["Grupo retirada"],
                    "Número do pedido de retirada": row["Número do pedido de retirada"],
                    "Tempo de retirada": row["Tempo de retirada_dt"],
                    "Tempo de sucesso": row["Tempo de sucesso_dt"],
                    "Status Shein": row["Status de retirada"],
                    "Valor Shein": row["Valor recebido"],
                    "Valor total do grupo": row["Valor total do grupo"],
                    "Situacao conciliacao": "Ainda pendente na Shein",
                }
            )

    bank_unmatched = []
    for idx, bank in enumerate(bank_rows):
        if idx not in matched_bank_indexes:
            bank_unmatched.append({**bank, "Situacao conciliacao": "Nao localizado na Shein"})
    return reconciled, shein_pending, bank_unmatched


if __name__ == "__main__":
    import json

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    shein_rows = load_shein_rows()
    bank_rows = load_bank_rows()
    shein_groups = summarize_shein(shein_rows)
    reconciled, shein_pending, bank_unmatched = reconcile(shein_rows, bank_rows)
    payload = {
        "shein_rows": shein_rows,
        "shein_groups": shein_groups,
        "bank_rows": bank_rows,
        "reconciled": reconciled,
        "shein_pending": shein_pending,
        "bank_unmatched": bank_unmatched,
    }

    def default(obj: Any) -> str:
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return str(obj)

    ANALYSIS_JSON.write_text(json.dumps(payload, default=default, ensure_ascii=False, indent=2), encoding="utf-8")
    print("shein_groups", len(shein_groups), "shein_rows", len(shein_rows), "bank_rows", len(bank_rows))
    print("reconciled", len(reconciled), "shein_pending", len(shein_pending), "bank_unmatched", len(bank_unmatched))
    print("shein_success_total", round(sum(g["Valor bem-sucedido"] for g in shein_groups), 2))
    print("shein_success_items_total", round(sum(r["Valor recebido"] for r in shein_rows if r["Status de retirada"] == "Retirada bem -sucedida"), 2))
    print("bank_total", round(sum(b["Credito"] for b in bank_rows), 2))
    for item in reconciled:
        print("OK", item["Número do pedido de retirada"], item["Valor Shein"], item["Data banco"].date(), item["Documento banco"])
    for item in shein_pending:
        print("PEND_SHEIN", item["Número do pedido de retirada"], item["Valor Shein"], item["Status Shein"], item["Situacao conciliacao"])
    for item in bank_unmatched:
        print("PEND_BANK", item["Data"], item["Documento"], item["Credito"])
